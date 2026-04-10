import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from eventos.models import Evento, Presupuesto
from cobros.models import Cobro, Pago
from inventario.models import Producto, MovimientoStock
from gastos.models import GastoFijo


# Shared styles
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
MONEY_FORMAT = '#,##0'
PCT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def safe_n(val):
    """Safely convert a value to float for Excel, defaults to 0."""
    try:
        if val is None: return 0.0
        return float(val)
    except:
        return 0.0

def safe_s(val):
    """Safely convert a value to string, handling None and FieldFiles."""
    if not val: return ""
    # If it's a FieldFile, return the name or URL if available
    if hasattr(val, 'url'):
        try:
            return val.url
        except:
            return str(val)
    return str(val)

def style_header(ws, row=1, cols=10):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def make_response(wb, filename):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class ExportEventosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Eventos'

        headers = ['ID', 'Nombre', 'Cliente', 'Fecha', 'Tipo', 'Pax', 'Lugar',
                   'Estado', 'Costo Total', 'Venta Total', 'Utilidad', 'Margen %']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        eventos = Evento.objects.all().order_by('-fecha')
        anio = request.query_params.get('anio')
        estado = request.query_params.get('estado')
        if anio:
            eventos = eventos.filter(fecha__year=int(anio))
        if estado:
            eventos = eventos.filter(estado=estado)

        for row, e in enumerate(eventos, 2):
            ws.cell(row=row, column=1, value=e.id)
            ws.cell(row=row, column=2, value=safe_s(e.nombre))
            ws.cell(row=row, column=3, value=safe_s(e.cliente))
            ws.cell(row=row, column=4, value=e.fecha.strftime('%d/%m/%Y') if e.fecha else "")
            ws.cell(row=row, column=5, value=e.get_tipo_evento_display())
            ws.cell(row=row, column=6, value=safe_n(e.pax))
            ws.cell(row=row, column=7, value=safe_s(e.lugar))
            ws.cell(row=row, column=8, value=e.get_estado_display())
            ws.cell(row=row, column=9, value=safe_n(e.costo_total)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=10, value=safe_n(e.venta_total)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=11, value=safe_n(e.utilidad)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=12, value=safe_n(e.margen))

        auto_width(ws)
        return make_response(wb, f'eventos_{datetime.now().strftime("%Y%m%d")}.xlsx')


class ExportPresupuestosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Presupuestos'

        headers = ['N° Presupuesto', 'Evento', 'Cliente', 'Estado', 'Subtotal',
                   'Descuento %', 'Total', 'Forma de Pago', 'Fecha Creación']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        presupuestos = Presupuesto.objects.select_related('evento').all()
        for row, p in enumerate(presupuestos, 2):
            ws.cell(row=row, column=1, value=safe_s(p.numero))
            ws.cell(row=row, column=2, value=safe_s(p.evento.nombre))
            ws.cell(row=row, column=3, value=safe_s(p.evento.cliente))
            ws.cell(row=row, column=4, value=p.get_estado_display())
            ws.cell(row=row, column=5, value=safe_n(p.subtotal)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=6, value=safe_n(p.descuento_pct))
            ws.cell(row=row, column=7, value=safe_n(p.total)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=8, value=p.get_forma_pago_display())
            ws.cell(row=row, column=9, value=p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else "")

        # Detail sheet
        ws2 = wb.create_sheet('Detalle Items')
        headers2 = ['N° Presupuesto', 'Descripción', 'Categoría', 'Cantidad',
                     'Costo Unit.', 'Venta Unit.', 'Total Costo', 'Total Venta', 'Margen %']
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=h)
        style_header(ws2, 1, len(headers2))

        row = 2
        for p in presupuestos.prefetch_related('items'):
            for item in p.items.all():
                ws2.cell(row=row, column=1, value=p.numero)
                ws2.cell(row=row, column=2, value=safe_s(item.descripcion))
                ws2.cell(row=row, column=3, value=item.get_categoria_display())
                ws2.cell(row=row, column=4, value=safe_n(item.cantidad))
                ws2.cell(row=row, column=5, value=safe_n(item.costo_unitario)).number_format = MONEY_FORMAT
                ws2.cell(row=row, column=6, value=safe_n(item.venta_unitario)).number_format = MONEY_FORMAT
                ws2.cell(row=row, column=7, value=safe_n(item.total_costo)).number_format = MONEY_FORMAT
                ws2.cell(row=row, column=8, value=safe_n(item.total_venta)).number_format = MONEY_FORMAT
                ws2.cell(row=row, column=9, value=safe_n(item.margen))
                row += 1

        auto_width(ws)
        auto_width(ws2)
        return make_response(wb, f'presupuestos_{datetime.now().strftime("%Y%m%d")}.xlsx')


class ExportCobrosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cobros'

        headers = ['ID', 'N° Presupuesto', 'Evento', 'Cliente', 'Monto Total',
                   'Monto Pagado', 'Saldo Pendiente', 'Estado', 'Fecha Envío']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        cobros = Cobro.objects.select_related('presupuesto__evento').prefetch_related('pagos').all()
        for row, c in enumerate(cobros, 2):
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=safe_s(c.presupuesto.numero))
            ws.cell(row=row, column=3, value=safe_s(c.presupuesto.evento.nombre))
            ws.cell(row=row, column=4, value=safe_s(c.presupuesto.evento.cliente))
            ws.cell(row=row, column=5, value=safe_n(c.monto_total)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=6, value=safe_n(c.monto_pagado)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=7, value=safe_n(c.saldo_pendiente)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=8, value=c.get_estado_display())
            ws.cell(row=row, column=9, value=c.fecha_envio.strftime('%d/%m/%Y') if c.fecha_envio else '')

        # Pagos sheet
        ws2 = wb.create_sheet('Pagos')
        headers2 = ['Cobro ID', 'N° Presupuesto', 'Monto Pago', 'Método', 'Fecha Pago', 'Comprobante']
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=h)
        style_header(ws2, 1, len(headers2))

        row = 2
        for pago in Pago.objects.select_related('cobro__presupuesto').all().order_by('-fecha_pago'):
            ws2.cell(row=row, column=1, value=pago.cobro.id)
            ws2.cell(row=row, column=2, value=safe_s(pago.cobro.presupuesto.numero))
            ws2.cell(row=row, column=3, value=safe_n(pago.monto)).number_format = MONEY_FORMAT
            ws2.cell(row=row, column=4, value=pago.get_metodo_pago_display())
            ws2.cell(row=row, column=5, value=pago.fecha_pago.strftime('%d/%m/%Y') if pago.fecha_pago else "")
            ws2.cell(row=row, column=6, value=safe_s(pago.comprobante))
            row += 1

        auto_width(ws)
        auto_width(ws2)
        return make_response(wb, f'cobros_pagos_{datetime.now().strftime("%Y%m%d")}.xlsx')


class ExportStockView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        headers = ['ID', 'Producto', 'Categoría', 'Unidad', 'Stock Actual',
                   'Stock Mínimo', 'Precio Compra', 'Precio Venta', 'Proveedor',
                   'Valor Inventario', 'Margen %', 'Stock Bajo']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        productos = Producto.objects.filter(activo=True).order_by('nombre')
        for row, p in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=p.id)
            ws.cell(row=row, column=2, value=safe_s(p.nombre))
            ws.cell(row=row, column=3, value=p.get_categoria_display())
            ws.cell(row=row, column=4, value=p.get_unidad_display())
            ws.cell(row=row, column=5, value=safe_n(p.stock_actual))
            ws.cell(row=row, column=6, value=safe_n(p.stock_minimo))
            ws.cell(row=row, column=7, value=safe_n(p.precio_compra)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=8, value=safe_n(p.precio_venta)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=9, value=safe_s(p.proveedor))
            ws.cell(row=row, column=10, value=safe_n(p.valor_inventario)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=11, value=safe_n(p.margen))
            ws.cell(row=row, column=12, value='⚠️ SÍ' if p.stock_bajo else 'OK')

        auto_width(ws)
        return make_response(wb, f'inventario_{datetime.now().strftime("%Y%m%d")}.xlsx')


class ExportGastosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Gastos Fijos'

        headers = ['ID', 'Nombre', 'Categoría', 'Monto', 'Vencimiento', 'Estado', 'Comprobante', 'Observaciones']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        mes = request.query_params.get('mes')
        anio = request.query_params.get('anio')
        
        gastos = GastoFijo.objects.all().order_by('-fecha_vencimiento')
        if mes: gastos = gastos.filter(mes=mes)
        if anio: gastos = gastos.filter(anio=anio)

        for row, g in enumerate(gastos, 2):
            ws.cell(row=row, column=1, value=g.id)
            ws.cell(row=row, column=2, value=safe_s(g.nombre))
            ws.cell(row=row, column=3, value=g.get_categoria_display())
            ws.cell(row=row, column=4, value=safe_n(g.monto)).number_format = MONEY_FORMAT
            ws.cell(row=row, column=5, value=g.fecha_vencimiento.strftime('%d/%m/%Y') if g.fecha_vencimiento else "")
            ws.cell(row=row, column=6, value=g.get_estado_display())
            ws.cell(row=row, column=7, value=safe_s(g.comprobante))
            ws.cell(row=row, column=8, value=safe_s(g.observaciones))

        auto_width(ws)
        return make_response(wb, f'gastos_{datetime.now().strftime("%Y%m%d")}.xlsx')


class ExportDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumen Dashboard'
        anio_val = request.query_params.get('anio', datetime.now().year)
        try:
            anio = int(anio_val)
        except:
            anio = datetime.now().year

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'GESTIÓN INTEGRAL - Resumen {anio}'
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='0F1923')

        # Monthly summary
        headers = ['Mes', 'Eventos', 'Ventas', 'Costos', 'Utilidad', 'Margen %']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))

        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        total_v, total_c = 0.0, 0.0
        for m in range(1, 13):
            row = m + 3
            eventos_count = Evento.objects.filter(fecha__year=anio, fecha__month=m).count()
            presupuestos = Presupuesto.objects.filter(evento__fecha__year=anio, evento__fecha__month=m)
            
            ventas = sum(safe_n(p.total) for p in presupuestos)
            costos = 0.0
            for p in presupuestos.prefetch_related('items'):
                costos += sum(safe_n(i.cantidad) * safe_n(i.costo_unitario) for i in p.items.all())
            
            util = ventas - costos
            margen = (util / ventas * 100) if ventas > 0 else 0.0
            total_v += ventas
            total_c += costos

            ws.cell(row=row, column=1, value=meses[m - 1])
            ws.cell(row=row, column=2, value=eventos_count)
            ws.cell(row=row, column=3, value=ventas).number_format = MONEY_FORMAT
            ws.cell(row=row, column=4, value=costos).number_format = MONEY_FORMAT
            ws.cell(row=row, column=5, value=util).number_format = MONEY_FORMAT
            ws.cell(row=row, column=6, value=round(margen, 1))

        # Totals row
        row = 16
        ws.cell(row=row, column=1, value='TOTALES').font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_v).number_format = MONEY_FORMAT
        ws.cell(row=row, column=4, value=total_c).number_format = MONEY_FORMAT
        ws.cell(row=row, column=5, value=total_v - total_c).number_format = MONEY_FORMAT
        tot_margen = ((total_v - total_c) / total_v * 100) if total_v > 0 else 0.0
        ws.cell(row=row, column=6, value=round(tot_margen, 1))
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)

        auto_width(ws)
        return make_response(wb, f'dashboard_resumen_{anio}.xlsx')
